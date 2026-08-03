from __future__ import annotations

from hashlib import sha256
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_to_tuple

from kms_agent.models import DocumentElement, SourceLocation


class ExcelExtractor:
    def extract(self, path: str | Path) -> list[DocumentElement]:
        source = Path(path)
        checksum = sha256(source.read_bytes()).hexdigest()
        document_id = checksum[:16]
        keep_vba = source.suffix.lower() == ".xlsm"
        workbook = load_workbook(source, data_only=False, read_only=False, keep_vba=keep_vba)
        cached = load_workbook(source, data_only=True, read_only=False, keep_vba=keep_vba)
        elements: list[DocumentElement] = []
        for sheet in workbook.worksheets:
            cached_sheet = cached[sheet.title]
            header_row, headers = self._headers(sheet)
            base_metadata = {
                "sheet": sheet.title,
                "sheet_state": sheet.sheet_state,
                "merged_ranges": [str(item) for item in sheet.merged_cells.ranges],
            }
            for row in sheet.iter_rows():
                populated = [cell for cell in row if cell.value not in (None, "") or cell.comment]
                if not populated:
                    continue
                start = populated[0].coordinate
                end = populated[-1].coordinate
                parts: list[str] = []
                for cell in populated:
                    value = cell.value
                    header = headers.get(cell.column) if cell.row != header_row else None
                    field = f"{header} ({cell.coordinate})" if header else cell.coordinate
                    if isinstance(value, str) and value.startswith("="):
                        cached_value = cached_sheet[cell.coordinate].value
                        formula_text = f"{field} formula {value}"
                        if cached_value is not None:
                            formula_text += f"; calculated value {cached_value}"
                        if cell.comment:
                            formula_text += f"; comment {cell.comment.text}"
                        elements.append(
                            self._element(
                                document_id,
                                source.name,
                                checksum,
                                sheet.title,
                                cell.coordinate,
                                "formula",
                                formula_text,
                                {**base_metadata, "formula": value, "cached_value": cached_value},
                            )
                        )
                    display = f"{field}={value}"
                    if cell.comment:
                        display += f" [comment: {cell.comment.text}]"
                    parts.append(display)
                elements.append(
                    self._element(
                        document_id,
                        source.name,
                        checksum,
                        sheet.title,
                        start if start == end else f"{start}:{end}",
                        "row",
                        " | ".join(parts),
                        base_metadata,
                    )
                )
            elements.extend(self._chart_elements(document_id, source.name, checksum, workbook, sheet, base_metadata))
            elements.extend(self._sheet_metadata(document_id, source.name, checksum, sheet, base_metadata))

        defined_names = [str(name) for name in workbook.defined_names.values()]
        if defined_names or keep_vba:
            text = "Workbook metadata"
            if defined_names:
                text += "\nDefined names: " + " | ".join(defined_names)
            if keep_vba:
                text += "\nContains VBA project; macros were not executed."
            elements.append(
                DocumentElement.create(
                    document_id=document_id,
                    document_title=source.name,
                    document_checksum=checksum,
                    location=SourceLocation(kind="xlsx", label="Workbook metadata"),
                    content_type="workbook_metadata",
                    text=text,
                    metadata={"contains_vba": keep_vba},
                )
            )
        workbook.close()
        cached.close()
        return elements

    @staticmethod
    def _headers(sheet):
        for row in sheet.iter_rows():
            populated = [cell for cell in row if cell.value not in (None, "")]
            if populated:
                return populated[0].row, {cell.column: str(cell.value) for cell in populated}
        return 0, {}

    def _chart_elements(self, document_id, title, checksum, workbook, sheet, base_metadata):
        elements = []
        for index, chart in enumerate(sheet._charts, start=1):
            xml = chart._write()
            texts = []
            references = []
            for node in xml.iter():
                local_name = node.tag.rsplit("}", 1)[-1]
                value = (node.text or "").strip()
                if not value:
                    continue
                if local_name == "t" and value not in texts:
                    texts.append(value)
                elif local_name == "f" and value not in references:
                    references.append(value)
            lines = []
            if texts:
                lines.append("Chart text: " + " | ".join(texts))
            for reference in references:
                resolved = self._resolve_reference(workbook, reference)
                lines.append(f"Chart source {reference}: {resolved}" if resolved else f"Chart source {reference}")
            if not lines:
                lines.append(f"Chart {index}")
            anchor = self._chart_anchor(chart)
            label = f"Chart {index}" + (f" at {anchor}" if anchor else "")
            elements.append(
                self._element(
                    document_id,
                    title,
                    checksum,
                    sheet.title,
                    label,
                    "chart",
                    "\n".join(lines),
                    {**base_metadata, "chart_index": index, "source_references": references},
                )
            )
        return elements

    @staticmethod
    def _resolve_reference(workbook, reference):
        try:
            sheet_name, bounds = range_to_tuple(reference)
            min_col, min_row, max_col, max_row = bounds
            source_sheet = workbook[sheet_name]
        except (ValueError, KeyError, TypeError):
            return ""
        values = []
        for row in source_sheet.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
            for cell in row:
                if cell.value not in (None, ""):
                    values.append(f"{cell.coordinate}={cell.value}")
        return " | ".join(values)

    @staticmethod
    def _chart_anchor(chart):
        try:
            return f"{get_column_letter(chart.anchor._from.col + 1)}{chart.anchor._from.row + 1}"
        except (AttributeError, TypeError):
            return ""

    def _sheet_metadata(self, document_id, title, checksum, sheet, base_metadata):
        details = []
        if sheet.tables:
            details.append("Tables: " + ", ".join(sheet.tables.keys()))
        if sheet._charts:
            details.append(f"Charts: {len(sheet._charts)}")
        validations = getattr(sheet.data_validations, "count", 0)
        if validations:
            details.append(f"Data validation rules: {validations}")
        conditional = len(sheet.conditional_formatting)
        if conditional:
            details.append(f"Conditional formatting ranges: {conditional}")
        if not details:
            return []
        return [
            self._element(
                document_id,
                title,
                checksum,
                sheet.title,
                "Sheet metadata",
                "sheet_metadata",
                "\n".join(details),
                base_metadata,
            )
        ]

    @staticmethod
    def _element(document_id, title, checksum, sheet, cells, content_type, text, metadata):
        location = SourceLocation(
            kind="xlsx",
            label=f"Sheet '{sheet}' | {cells}",
            coordinates={"sheet": sheet, "cell_range": cells},
        )
        return DocumentElement.create(
            document_id=document_id,
            document_title=title,
            document_checksum=checksum,
            location=location,
            content_type=content_type,
            text=text,
            metadata=dict(metadata),
        )
