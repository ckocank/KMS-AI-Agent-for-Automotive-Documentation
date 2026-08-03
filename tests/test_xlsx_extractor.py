import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.comments import Comment

from tests.helpers import SRC  # noqa: F401
from kms_agent.extractors.xlsx import ExcelExtractor


class ExcelExtractorTests(unittest.TestCase):
    def test_extracts_formula_comment_and_hidden_sheet(self):
        with tempfile.TemporaryDirectory() as folder:
            path = Path(folder) / "coverage.xlsx"
            book = Workbook()
            sheet = book.active
            sheet.title = "Coverage"
            sheet.append(["Passed", "Total", "Rate"])
            sheet.append([90, 100, "=A2/B2"])
            sheet["C2"].comment = Comment("Target >= 90%", "QA")
            hidden = book.create_sheet("Internal")
            hidden.sheet_state = "hidden"
            hidden["A1"] = "ASIL-D review"
            book.save(path)

            elements = ExcelExtractor().extract(path)

        text = "\n".join(item.text for item in elements)
        labels = [item.location.label for item in elements]
        self.assertIn("=A2/B2", text)
        self.assertIn("Target >= 90%", text)
        self.assertTrue(any("Sheet 'Coverage'" in label and "C2" in label for label in labels))
        self.assertTrue(any(item.metadata.get("sheet_state") == "hidden" for item in elements))

    def test_extracts_chart_series_and_source_values(self):
        with tempfile.TemporaryDirectory() as folder:
            path = Path(folder) / "validation.xlsx"
            book = Workbook()
            sheet = book.active
            sheet.title = "Results"
            sheet.append(["Build", "Passed"])
            sheet.append(["B100", 92])
            sheet.append(["B101", 97])
            chart = BarChart()
            chart.title = "Validation Pass Count"
            chart.add_data(Reference(sheet, min_col=2, min_row=1, max_row=3), titles_from_data=True)
            chart.set_categories(Reference(sheet, min_col=1, min_row=2, max_row=3))
            sheet.add_chart(chart, "E2")
            book.save(path)

            elements = ExcelExtractor().extract(path)

        chart_text = "\n".join(item.text for item in elements if item.content_type == "chart")
        self.assertIn("Validation Pass Count", chart_text)
        self.assertIn("B100", chart_text)
        self.assertIn("97", chart_text)


if __name__ == "__main__":
    unittest.main()
